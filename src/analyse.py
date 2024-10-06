from decouple import config
from preprocessor import xml_files_reader
import os, time, re
import numpy as np
import pandas as pd
from system import ManagedSystem
from report import Report, summaryHeaders, reportHeaders
from constants import (
    attribute_string_operators,
    attribute_string_keywords,
    system_event_rule,
    another_situation_reference,
    pdt_ignore,
    maint_window,
    INSTANA_SEV_WARNING,
    INSTANA_SEV_CRITICAL,
    INSTANA_SEV_INFO,
    SUPPORTED,
    NOT_SUPPORTED,
    CUSTOM,
    UNKNOWN,
)
from shared import (
    logger,
    MIGRATION_PATH,
    NAME,
    itm2instana_mappings_df,
    metric_mappings_df,
    product_codes_df,
)

# Import necessary spreadsheet style classes
from openpyxl.utils import get_column_letter

# =================================
# Setup Migration Tool environment:
# =================================
# - Define Analyser constants
OUTPUT_ANALYSIS_PATH = MIGRATION_PATH + "analysis" + os.sep + NAME + os.sep
ANALYSIS_SEED_DATA_PATH = OUTPUT_ANALYSIS_PATH + "seed_data" + os.sep
ANALYSIS_REPORT_FILENAME = OUTPUT_ANALYSIS_PATH + NAME + "_report.xlsx"
ANALYSIS_SUMMARY_FILENAME = OUTPUT_ANALYSIS_PATH + NAME + "_summary.md"
# Create folder paths if they do not already exist - makedirs succeeds even if directory exists.
os.makedirs(OUTPUT_ANALYSIS_PATH, exist_ok=True)
os.makedirs(ANALYSIS_SEED_DATA_PATH, exist_ok=True)
# Start logger
start_time = time.time()
logger.info("Analysis started...")


# =================================
# Pre-Processing
# =================================


# Return seed data filename for a given folder
def getAnalysisSeedDataFilename(folderId):
    # prodSupported = is_productCode_Supported(folderId)
    return ANALYSIS_SEED_DATA_PATH + "attributeGroup_items_" + folderId + ".csv"


# Return ITM attributes list for a given folder
def loadAttributesForFolder(folderId):
    lines = None
    analysisSeedDataFilename = getAnalysisSeedDataFilename(folderId)
    logger.debug("loadAttributesForFolder:" + analysisSeedDataFilename)
    if os.path.isfile(analysisSeedDataFilename):
        try:
            with open(analysisSeedDataFilename) as file:
                lines = [line.rstrip() for line in file]
        except:
            logger.error(
                "Folder Attribute Mappings not found in seed data i.e. ",
                analysisSeedDataFilename,
            )
    return lines


# Create filtered list of attributes used in Situations, by folder, and check if metric mapping exists
def create_profiling_on_pdt_formula_to_attributeGroup_MetricItem(sits_list, folderId):
    logger.debug("Writing list of attributes found in Active Situations")
    attribute_item_list = []
    errors = []
    supported = []
    notSupported = []
    for sit in sits_list:
        migrate = sit.get_autostart()
        if migrate:
            pdt_formula = sit.get_pdt()
            # split pdt formula based on spaces ' '
            list_items = pdt_formula.split()
            iterator = 0
            for iterator in range(len(list_items)):
                item = list_items[iterator]
                # find attribute_group.attribute_Item in seed data metric mappings
                if "." in item and (
                    list_items[iterator - 1] in attribute_string_keywords
                ):
                    if item not in attribute_item_list:
                        attribute_item_list.append(item)
                iterator += 1
                # Write summary list of ITM attributes found in Situation folder data
                with open(getAnalysisSeedDataFilename(folderId), "w") as tfile:
                    tfile.write("\n".join(attribute_item_list))
    # Check if metric mapping exists
    for attrItem in attribute_item_list:
        attr = attrItem.split(".")
        metric_mapping = metric_mappings_df[
            (metric_mappings_df["ITM_Attribute_Group"] == attr[0])
            & (metric_mappings_df["ITM_Attribute_Item"] == attr[1])
        ]
        if len(metric_mapping) == 0:
            errors.append(attrItem)
        else:
            metricSupported = get_attribute_value_from_df(metric_mapping, "Supported")
            if metricSupported != "SUPPORTED":
                notSupported.append(attrItem)
            else:
                supported.append(attrItem)
    mappingMessage = str(len(attribute_item_list)) + " unique metric mappings found"
    notSupportedCount = len(np.unique(notSupported))
    supportedCount = len(np.unique(supported))
    if notSupportedCount > 0:
        isAre = "is" if notSupportedCount == 1 else "are"
        mappingMessage += (
            ", of which "
            + str(notSupportedCount)
            + " metrics "
            + isAre
            + " not supported and "
            + str(len(supported))
            + " supported."
        )
    logger.info(mappingMessage)
    for attrItem in sorted(notSupported):
        logger.warning("Metric not supported: " + folderId + " -> " + attrItem)
    for supportedMetric in sorted(supported):
        logger.info("Metric mapping found and is supported: " + supportedMetric)
    if len(errors) > 0:
        logger.info(
            str(len(errors)) + " Metric Mappings not found in folder: " + folderId
        )
        for attrItem in sorted(errors):
            logger.warning("No Metric Mapping found: " + folderId + " -> " + attrItem)
    todo = len(attribute_item_list) - notSupportedCount - supportedCount
    if todo > 0:
        logger.error("TODO!!! " + str(todo) + " metric mappings not yet mapped")


# =================================
# Analysis
# =================================


# Return name associated with Product code using ITMProductCodes.csv
def get_productCode_Name(productCode):
    productCodeName = None
    try:
        productCodeName = product_codes_df[
            (product_codes_df["ProductCode"] == productCode)
        ]["ProductName"].item()
    except:
        logger.warning("Agent Folder not recognised: " + productCode)
    return productCodeName


# Return name associated with Product code using ITMProductCodes.csv
def is_productCode_Supported(productCode):
    supported = None
    # special case for Local Time attributes which may be associated with multiple products
    if productCode == "*": supported = "SUPPORTED"  
    try:
        supported = product_codes_df[(product_codes_df["ProductCode"] == productCode)][
            "Supported"
        ].item()
    except:
        logger.warning("Agent Folder not recognised: " + productCode)
    return supported


def get_metricMappings_for_ITM_attributegroup_and_attributeItem(
    attribute_group, attribute_item
):
    metric_mapping = None
    try:
        metric_mapping = metric_mappings_df[
            (metric_mappings_df["ITM_Attribute_Group"] == attribute_group)
            & (metric_mappings_df["ITM_Attribute_Item"] == attribute_item)
        ]
    except:
        logger.warning("Metric Mappings not found in seed data i.e. metricMapping.csv")
    return metric_mapping


def get_attribute_value_from_df(attribute_df, attribute_name):
    attribute_value = None
    try:
        # set to None, if record is present but value is empty in dataframe
        if attribute_df[attribute_name].item() and not isNaN(
            attribute_df[attribute_name].item()
        ):
            attribute_value = attribute_df[attribute_name].item()
    except:
        attribute_value = None
    return attribute_value


def analyseSituationPDT(sitReport, folderIdSupported):
    pdt_formula = sitReport.pdt
    name = sitReport.sitName
    logger.debug(name + ": " + pdt_formula)
    rules = []  # rules object to be added to the current event
    rangeMetrics = []
    supported = []
    missingMetric = []
    query = None
    entity_type = None
    mappingConfidence = []

    pdt_array = pdt_formula.split()
    if another_situation_reference in pdt_array:
        logger.debug(
            "The situation referencing another situation not supported. " + pdt_formula
        )
        sitReport.hasSitRef = True

    # check if pdt formula items are found in system event rule array
    is_system_rule = [True for item in pdt_array if item in system_event_rule]
    sitReport.isSystem = len(is_system_rule) > 0

    def get_values_for_single_condition(list_items_in_single_pdt):
        iterator = 0
        count = 0
        rule_metric_mappings_df = None
        condition_operator = None
        operand = None
        query_value = None
        for iterator in range(len(list_items_in_single_pdt)):
            item = list_items_in_single_pdt[iterator]
            # *IF *VALUE Queue_Statistics.Queue_Name *EQ Q289P.PROD.SA.SAST.REQUEST *AND *VALUE Queue_Statistics.Current_Depth *GE 1000
            #
            if item == "*IF" or item in attribute_string_keywords:
                keyword_string = item
                continue  # skipping

            # find attribute_group.attribute_Item, then split and retrieve metric mappings from seed data
            if "." in item and (
                list_items_in_single_pdt[iterator - 1] in attribute_string_keywords
            ):
                items = item.split(".")
                rule_metric_mappings_df = (
                    get_metricMappings_for_ITM_attributegroup_and_attributeItem(
                        attribute_group=items[0], attribute_item=items[1]
                    )
                )
                if rule_metric_mappings_df.empty and folderIdSupported == "SUPPORTED":
                    missingMetric.append(item)

            # check if there is an *EQ keyword, fetch its mapping value from seed data
            elif item in attribute_string_operators:
                condition_operator = get_keyword_mapping(item)
            elif (
                item
                and list_items_in_single_pdt[iterator - 1] in attribute_string_operators
            ):
                operand = item.strip("'\"")

            iterator += 1

        metricSupported = get_attribute_value_from_df(
            rule_metric_mappings_df, "Supported"
        )
        mappingConfidence.append(metricSupported)
        entity_type = get_attribute_value_from_df(rule_metric_mappings_df, "entityType")
        metric_name = get_attribute_value_from_df(rule_metric_mappings_df, "metricName")
        aggregation = get_attribute_value_from_df(
            rule_metric_mappings_df, "aggregation"
        )
        query_value = get_attribute_value_from_df(rule_metric_mappings_df, "query")
        metric_pattern_prefix = get_attribute_value_from_df(
            rule_metric_mappings_df, "metricPattern.prefix"
        )
        metric_pattern_postfix = get_attribute_value_from_df(
            rule_metric_mappings_df, "metricPattern.postfix"
        )
        # if metric_pattern_prefix or metric_pattern_postfix and sitReport.hasMetricPattern != True:
        sitReport.hasMetricPattern = (
            metric_pattern_prefix != None or metric_pattern_postfix != None
        )

        conversion_rule = get_attribute_value_from_df(
            rule_metric_mappings_df, "conversionRule"
        )

        sitReport.hasMaintWindow = conversion_rule == maint_window
        
        return (
            entity_type,
            metric_name,
            metric_pattern_prefix,
            metric_pattern_postfix,
            aggregation,
            query_value,
            condition_operator,
            operand,
            metricSupported,
            conversion_rule
        )

    def handle_single_condition(condition, query_operator, query, rules, entity_type):
        logger.debug(condition)
        # split pdt formula based on spaces ' '
        list_items_in_single_pdt = condition.split()
        (
            rule_entity_type,
            metric_name,
            metric_pattern_prefix,
            metric_pattern_postfix,
            aggregation,
            query_value,
            condition_operator,
            operand,
            metricSupported,
            conversion_rule
        ) = get_values_for_single_condition(list_items_in_single_pdt)

        supported.append(metricSupported)

        if entity_type is None:
            entity_type = rule_entity_type
        if (
            (
                metric_name is None
                and (metric_pattern_prefix is not None or query_value is not None)
            )
            or (
                metric_pattern_prefix is None
                and (metric_name is not None or query_value is not None)
            )
            or (
                query_value is None
                and (metric_name is not None or metric_pattern_prefix is not None)
            )
        ):
            sitReport.hasMapping = True
        else:
            sitReport.hasMapping = False

        if (
            rule_entity_type is not None
            and metric_name is None
            and metric_pattern_prefix is None
            and metric_pattern_postfix is None
            and query_value is not None
        ):
            if query is None:
                query = query_value + ":'" + operand + "'"
            elif query is not None:
                query = query + query_operator + query_value + ":'" + operand + "'"
        elif rule_entity_type is not None and metric_name is not None:
            duplicate_metric_rule = [r for r in rules if r["metricName"] == metric_name]
            if (
                duplicate_metric_rule
                and duplicate_metric_rule[0].get("conditionOperator").startswith(">")
                and condition_operator.startswith("<")
            ) or (
                duplicate_metric_rule
                and duplicate_metric_rule[0].get("conditionOperator").startswith("<")
                and condition_operator.startswith(">")
            ):
                rangeMetrics.append(metric_name)
                logger.debug(
                    "A RANGE with a duplicate metricName was found in PDT formula. Skipped adding a new rule into rules array."
                )
            else:
                rule = {
                    "ruleType": "threshold",
                    "metricName": metric_name,
                    "metricPattern": None,
                    "rollup": 0,
                    "window": 1000,
                    "aggregation": aggregation,
                    "conditionOperator": condition_operator,
                    "conditionValue": operand,
                }
                rules.append(rule)
        return query, rules, entity_type

    def handle_conditions(
        pdt_formula,
        query,
        rules,
        entity_type,
        query_operator=" OR ",
    ):
        if pdt_formula not in pdt_ignore and pdt_formula:
            pdt_sub_array = pdt_formula.split()
            if "(" in pdt_sub_array or ")" in pdt_sub_array:
                logger.debug("handling brackets, splitting ")
                pdt_regex_array = re.split(r"\(|\)", pdt_formula)
                logger.debug(pdt_regex_array)
                for item in pdt_regex_array:
                    query, rules, entity_type = handle_conditions(
                        item.strip(), query, rules, entity_type
                    )
            elif "*OR" in pdt_sub_array:
                query_operator = " OR "
                logger.debug("handling *ORs")
                pdt_ors = pdt_formula.split("*OR")
                logger.debug(pdt_ors)
                for item in pdt_ors:
                    query, rules, entity_type = handle_conditions(
                        item.strip(), query, rules, entity_type
                    )
            elif "*AND" in pdt_sub_array:
                query_operator = " AND "
                logger.debug("Handling *ANDs")
                pdt_ands = pdt_formula.split("*AND")
                logger.debug(pdt_ands)
                for item in pdt_ands:
                    query, rules, entity_type = handle_conditions(
                        item.strip(), query, rules, entity_type
                    )
            else:
                logger.debug("handling single condition() " + pdt_formula)
                query, rules, entity_type = handle_single_condition(
                    pdt_formula, query_operator, query, rules, entity_type
                )

        return query, rules, entity_type

    query, rules, entity_type = handle_conditions(
        pdt_formula,
        query,
        rules,
        entity_type,
    )

    sitReport.hasRange = len(rangeMetrics) > 0
    sitReport.supported = supported.count("SUPPORTED")
    sitReport.notSupported = supported.count("NOT_SUPPORTED")
    sitReport.todo = supported.count("TODO")
    sitReport.missing = len(missingMetric)
    sitReport.mappingScore = len(mappingConfidence) - mappingConfidence.count("SUPPORTED")
    return sitReport


def get_keyword_mapping(itm_keyword):
    keyword_mapping = None
    if itm_keyword != None:
        try:
            # TODO: pass dynamic columns names from dataframe rather than hard coded values
            keyword_mapping = itm2instana_mappings_df[
                itm2instana_mappings_df["ITM_XML"] == itm_keyword
            ]["Instana_JSON_value"].item()
        except:
            logger.warning(
                "Keyword Mappings not found in seed data i.e. itm2instanaMapping.csv : "
                + str(itm_keyword)
            )
    return keyword_mapping


# =================================
# Spreadsheet report utilities
# =================================
def isNaN(num):
    return num != num


def cleanValueForSpreadsheetDisplay(val):
    # handle None and (None,) and (nan,)
    if isNaN(val[0]) or val[0] is None:
        return "UNKNOWN"
    elif val != None and not len(val) == val.count(None):
        return val[0]
    else:
        return val


def getPercentageValue(count, total):
    return 0 if count == 0 or total == 0 else round(count / total, 2)


def updateSummaryCount(counter, val):
    if val != None and val != "*NONE" and val != 0:
        counter.append(val)
    return counter


def getSummaryReportForFolder(summary, sitReports):
    eifs = []
    distributions = []
    commands = []
    ranges = []
    mappings = []
    percentages = []
    changes = []
    histRules = []
    counts = []
    systemEvents = []
    maintWindows = []
    deltas = []
    metricPatterns = []
    untilSits = []
    advises = []
    sevInfos = []
    sevCrits = []
    sevWarns = []
    sevOthers = []
    mappingsSupportedCount = 0
    mappingsNotSupportedCount = 0
    mappingsTodoCount = 0
    missingMappingCount = 0
    mappingScore = 0
    for r in sitReports:
        eifs = updateSummaryCount(eifs, r.hasEIF)
        commands = updateSummaryCount(commands, r.command)
        distributions = updateSummaryCount(distributions, r.hasDistribution)
        ranges = updateSummaryCount(ranges, r.hasRange)
        mappings = updateSummaryCount(mappings, r.hasMapping)
        percentages = updateSummaryCount(percentages, r.hasPercentChange)
        changes = updateSummaryCount(changes, r.hasChange)
        histRules = updateSummaryCount(histRules, r.hasHistRule)
        counts = updateSummaryCount(counts, r.pdtCount)
        systemEvents = updateSummaryCount(systemEvents, r.isSystem)
        maintWindows = updateSummaryCount(maintWindows, r.hasMaintWindow)
        deltas = updateSummaryCount(deltas, r.hasDelta)
        metricPatterns = updateSummaryCount(metricPatterns, r.hasMetricPattern)
        untilSits = updateSummaryCount(untilSits, r.hasUntilSit)
        advises = updateSummaryCount(advises, r.hasCustomAdvise)
        sevInfos = updateSummaryCount(sevInfos, r.sevInfo)
        sevOthers = updateSummaryCount(sevOthers, r.sevOther)
        sevCrits = updateSummaryCount(sevCrits, r.sevCritical)
        sevWarns = updateSummaryCount(sevWarns, r.sevWarning)
        if r.supported > 0:
            mappingsSupportedCount += 1
        if r.notSupported > 0:
            mappingsNotSupportedCount += 1
        if r.todo > 0:
            mappingsTodoCount += 1
        if r.missing > 0:
            missingMappingCount += 1
        if r.mappingScore == 0:
            mappingScore += 1

    summary.folderIdSupported = cleanValueForSpreadsheetDisplay(
        summary.folderIdSupported
    )
    summary.active = len(sitReports)
    summary.hasDistribution = len(distributions)
    summary.hasEIF = len(eifs)
    summary.command = len(commands)
    summary.hasRange = len(ranges)
    summary.hasMapping = len(mappings)
    summary.hasPercentChange = len(percentages)
    summary.hasChange = len(changes)
    summary.hasHistRule = len(histRules)
    summary.pdtCount = len(counts)
    summary.isSystem = len(systemEvents)
    summary.hasMaintWindow = len(maintWindows)
    summary.hasDelta = len(deltas)
    summary.hasMetricPattern = len(metricPatterns)
    summary.hasUntilSit = len(untilSits)
    summary.advise = len(advises)
    summary.sevInfo = len(sevInfos)
    summary.sevOther = len(sevOthers)
    summary.sevCritical = len(sevCrits)
    summary.sevWarning = len(sevWarns)
    summary.supported = mappingsSupportedCount
    summary.notSupported = mappingsNotSupportedCount
    summary.todo = mappingsTodoCount
    summary.missing = missingMappingCount
    summary.mappingScore = mappingScore

    logger.debug("Folder Summary Report:", vars(summary))
    summary.printReport()
    return summary.getSummaryReportArray()


def getSummaryReportTotalsSupportedOnly(summaryRows):
    totCounter = [0 for i in range(len(summaryRows[0]))]
    totCounter[0] = "TOTAL SUPPORTED"
    totCounter[1] = len(summaryRows)
    totCounter[2] = ""
    totPercentages = ["" for i in range(len(summaryRows[0]))]
    totPercentages[0] = "%"
    for row in summaryRows:
        for index, val in enumerate(row):
            if index > 2:
                if isinstance(val, int):
                    totCounter[index] += val
                elif can_convert_to_int(val):
                    totCounter[index] += int(val)
                else:
                    totCounter[index] = 0

    totalActiveSituations = totCounter[4]
    totalSituations = totCounter[3]
    totPercentages[4] = getPercentageValue(totalActiveSituations, totalSituations)
    totPercentages[5] = getPercentageValue(totCounter[5], totalSituations)
    for index, val in enumerate(totCounter):
        if index > 5:
            totPercentages[index] = getPercentageValue(
                totCounter[index], totalActiveSituations
            )
    return [totCounter, totPercentages]


def getSummaryReportTotals(summaryRows):
    totCounter = [0 for i in range(len(summaryRows[0]))]
    totCounter[0] = "TOTAL"
    totCounter[1] = len(summaryRows)
    totCounter[2] = ""
    totPercentages = ["" for i in range(len(summaryRows[0]))]
    totPercentages[0] = "%"
    for row in summaryRows:
        for index, val in enumerate(row):
            if index > 2:
                if isinstance(val, int):
                    totCounter[index] += val
                elif can_convert_to_int(val):
                    totCounter[index] += int(val)
                else:
                    totCounter[index] = 0

    totalActiveSituations = totCounter[4]
    totalSituations = totCounter[3]
    totPercentages[4] = getPercentageValue(totalActiveSituations, totalSituations)
    totPercentages[5] = getPercentageValue(totCounter[5], totalSituations)
    for index, val in enumerate(totCounter):
        if index > 5:
            totPercentages[index] = getPercentageValue(
                totCounter[index], totalActiveSituations
            )
    
    return [totCounter, totPercentages]


def analyseFolderXML(sits_list, attributeList, folderId, prodName, prodSupported, systems, listSystemsCSV):
    summaryReport = Report(
        folderId=folderId,
        agent=prodName,
        total=len(sits_list),
        folderIdSupported=prodSupported,
    )
    reports = []

    # init first row
    initRow = []
    if attributeList != None and len(attributeList) > 0:
        row = np.concatenate((reportHeaders, attributeList))
    else:
        row = reportHeaders
    for i in range(len(row)):
        initRow.append("")
    rows = [initRow]

    updatedSystems = systems
    for situation in sits_list:
        if situation.get_autostart():
            pdt = situation.pdt
            sev = situation.get_severity()            
            sevMap = get_keyword_mapping(sev)
            distribution = situation.get_distribution()  
            distMSLs = 0          
            if distribution != None:
                distMSL = distribution.split(",")
                distMSLs = len(distMSL)     
                updatedSystems = updateManagedSystemsWithSituation(situation.sit_name, distMSL, updatedSystems, listSystemsCSV)                
                        
            customAdvise = (
                situation.advise != None
                and situation.advise != "*NONE"
                and not (situation.advise.startswith('ADVICE("k'))
            )
            # initialise Situation report
            report = Report(
                # supported=supported,
                folderId=folderId,
                sitName=situation.sit_name,
                pdt=pdt,
                severity=sev,
                sevCritical=sevMap == INSTANA_SEV_CRITICAL,
                sevWarning=sevMap == INSTANA_SEV_WARNING,
                sevInfo=sevMap == INSTANA_SEV_INFO,
                sevOther=sevMap != INSTANA_SEV_CRITICAL
                and sevMap != INSTANA_SEV_WARNING
                and sevMap != INSTANA_SEV_INFO,
                interval=situation.reev_time,
                command=situation.cmd,
                hasCustomAdvise=customAdvise,
                advise=situation.advise,
                description=situation.text,
                hasEIF=situation.map,
                hasDistribution=distribution,                
                pdtCount=pdt.count("*COUNT"),
                hasDelta=pdt.count("*DELTA"),
                hasHistRule=pdt.count("*HISTRULE"),
                hasPercentChange=pdt.count("*PCTCHANGE"),
                hasChange=pdt.count("*CHANGE"),
                hasUntilSit=pdt.count("*UNTIL ( *SIT"),
                map=situation.map,
                distribution=distribution,
                distCount=distMSLs,
                andCount=pdt.count("*AND"),
                orCount=pdt.count("*OR"),
            )
            # Update Situation report with results from further analysis
            report = analyseSituationPDT(
                sitReport=report, folderIdSupported=summaryReport.folderIdSupported
            )
            r = report.getSituationReportArray()

            for attr in attributeList:
                r.append(situation.pdt.count(attr))

            rows.append(r)
            reports.append(report)

    summary = getSummaryReportForFolder(summaryReport, reports)
    # remove initial dummy row if report data present
    if len(rows) > 1:
        rows.pop(0)
    return attributeList, rows, summary, updatedSystems


def formatSummarySheet(summarySheet):
    # cell ranges
    title_row = "1"
    total_row = summarySheet.max_row - 1
    percent_row = summarySheet.max_row
    value_cells = "B2:{col}{row}".format(
        col=get_column_letter(summarySheet.max_column), row=summarySheet.max_row
    )
    index_column = "A"
    active_rows = "E2:E{row}".format(
        # col=get_column_letter(summarySheet.max_column),
        row=summarySheet.max_row
        - 2
    )

    # set first column width (Agent name)
    summarySheet.column_dimensions[index_column].width = 45

    for row in summarySheet[value_cells]:
        for cell in row:
            if cell.value == SUPPORTED:
                cell.style = "Good"
            elif cell.value == NOT_SUPPORTED or cell.value == UNKNOWN:
                cell.style = "Bad"
            elif cell.value == CUSTOM:
                cell.style = "Neutral"
            else:
                cell.style = "Normal"
    for row in summarySheet[active_rows]:
        for cell in row:
            if cell.value == 0:
                cell.style = "Check Cell"
    # style header line last, so that headline style wins in cell A1
    for cell in summarySheet[title_row]:
        cell.style = "Headline 2"
    for cell in summarySheet[total_row]:
        cell.style = "Total"
    for cell in summarySheet[percent_row]:
        cell.style = "Percent"


def can_convert_to_int(string):
    # ignore None and (None,)
    if string is not None and not (len(string) == string.count(None)):
        try:
            int(string)
            return True
        except ValueError:
            return False



# load ITM Situations root folder and analyse each folder XML files
def analyseXML(inputDirectoryPath, group_names, host_names, systems, listSystemsCSV, allQManagers):
    logger.info("Loading ITM Situations found in " + inputDirectoryPath)
    iterator = 0
    if os.path.isdir(inputDirectoryPath) and os.path.exists(inputDirectoryPath):
        # writer = pd.ExcelWriter(ANALYSIS_REPORT_FILENAME, engine='xlsxwriter', options={'strings_to_numbers': True})
        with pd.ExcelWriter(ANALYSIS_REPORT_FILENAME) as writer:            
            summaryDFs = []
            dataDFs = {}
            prodsSupported = []
            prodsNotSupported = []
            systemSituations = systems
            folderAttributes = {}
            for folder, subs, files in os.walk(inputDirectoryPath):
                subFolders = os.path.split(folder)
                folderId = subFolders[len(subFolders) - 1]
                if len(files) > 0 and len(subs) == 0:
                    # Load ITM Situations by folder
                    sits_list = xml_files_reader(folder)
                    prodName = get_productCode_Name(folderId)
                    prodSupported = is_productCode_Supported(folderId)
                    if prodSupported == "SUPPORTED":
                        prodsSupported.append([folderId, prodName])
                    else: 
                        prodsNotSupported.append([folderId, prodName])
                    logger.info(
                        "Analysing "
                        + str(len(sits_list))
                        + " Situation files in folder: "
                        + folderId
                    )
                    if prodSupported is None or prodSupported != "SUPPORTED":
                        logger.error(folderId + " is NOT SUPPORTED")
                    create_profiling_on_pdt_formula_to_attributeGroup_MetricItem(
                        sits_list, folderId
                    )
                    attributeList = loadAttributesForFolder(folderId)                    

                    folderReport = analyseFolderXML(
                        sits_list, attributeList, folderId, prodName, prodSupported, systemSituations, listSystemsCSV
                    )
                    if folderReport[0] != None:
                        reportHeader = np.concatenate((reportHeaders, folderReport[0]))
                    else:
                        reportHeader = reportHeaders
                    logger.info("Generating " + folderId + " report")
                    dataDFs[folderId] = [folderReport[1], reportHeader]
                    summaryDFs.append(folderReport[2])
                    systemSituations = folderReport[3]

            totalsummaryRows = getSummaryReportTotals(summaryDFs)
            
            totalsummaryRowsSupported = getSummaryReportTotalsSupportedOnly(summaryDFs)
            totCounter = totalsummaryRows[0]
            totPercentages = totalsummaryRows[1]
            totalActiveSituations = totCounter[4]
            totalSituations = totCounter[3]
            summaryDFs.append(totCounter)
            summaryDFs.append(totPercentages)
            
            # Write Summary Tab as first sheet, freeze top row and first 3 columns
            if len(summaryDFs) > 0:
                pd.DataFrame(summaryDFs).to_excel(
                    writer,
                    sheet_name="Summary",
                    header=summaryHeaders,
                    index=None,
                    freeze_panes=(1, 3),
                )
            # Format Summary sheet cell values
            summarySheet = writer.sheets["Summary"]
            formatSummarySheet(summarySheet)

            # Managed Systems data tabs
            if len(group_names) > 0: 
                pd.DataFrame(group_names).to_excel(
                        writer, sheet_name="Managed System List", header=["MSL GroupName"], index=None,
                    ) 
            hostsystems = []
            for host in host_names:
                sys = systemSituations[host]
                agentList = ','.join(map(str, sys.agents)) if len(sys.agents) > 0 else ''
                groupList = ','.join(map(str, sys.groups)) if len(sys.groups) > 0 else ''
                situationList = ','.join(map(str, sys.situations)) if len(sys.situations) > 0 else ''
                hostsystem = [host, len(sys.agents), agentList, len(sys.groups), groupList, len(sys.situations), situationList]
                hostsystems.append(hostsystem)

            if len(hostsystems) > 0: 
                pd.DataFrame(hostsystems).to_excel(
                        writer, sheet_name="Systems", header=["System", "Agent #", "Agents", "MSL #", "MSLs", "Situations #", "Situations"], index=None,
                    )
            
            if len(allQManagers) > 0: 
                pd.DataFrame(allQManagers).to_excel(
                        writer, sheet_name="MQMgrs", header=["MQ Manager"], index=None,
                    )

            # MAPPING ANALYSIS
            mapHeaders,mapRows = analyseMappings(folderId)
            pd.DataFrame(mapRows).to_excel(
                    writer, sheet_name="MAPPING", header=mapHeaders, index=None, freeze_panes=(1, 1),
                )
            # Write folder tabs after Summary sheet
            for key, value in list(dataDFs.items()):
                pd.DataFrame(value[0]).to_excel(
                    writer,
                    sheet_name=key,
                    index_label="Index",
                    header=value[1],
                )
                sh = writer.sheets[key]
                prodSupported = is_productCode_Supported(key)
                if prodSupported == "SUPPORTED":
                    sh.sheet_properties.tabColor = "C6EFCE"
                elif prodSupported == "CUSTOM":
                    sh.sheet_properties.tabColor = "FFEB9C"
                else:
                    sh.sheet_properties.tabColor = "FFC7CE"

            logger.info("Analysis report saved as " + ANALYSIS_REPORT_FILENAME)
            
            with open(ANALYSIS_SUMMARY_FILENAME, "w") as summaryFile:
                dataName = config("MIGRATION_NAME")
                
                writeSummaryReport(summaryFile, dataName, len(summaryDFs), totalSituations, totalActiveSituations, summaryHeaders, totCounter, totPercentages, len(group_names), len(host_names))
                sentence1 = "The Migration Tool may be able to migrate attributes from the following " + str(len(prodsSupported)) + " ITM agent types, "
                sentence2 = "(i.e. some but not all ITM attributes have mappings):\n"
                printSummaryList(summaryFile, sentence1 + sentence2, prodsSupported)
                
                printSummaryList(summaryFile, "The following " + str(len(prodsNotSupported)) + " ITM agent types are not supported by the Migration Tool:", prodsNotSupported)
                summaryFile.close()           

    else:
        logger.error(
            "Please provide a valid directory path in 'ITM_FOLDER' environment variable."
        )
        return None
    logger.info(
        "Total execution time: " + str(round(time.time() - start_time, 3)) + " seconds"
    )
    return iterator
    

# Create short summary paragraph of analysis results, e.g.:
# Client data has 22 folders with a total of 858 Situations of which 503 are active for migration. 
# 
def writeSummaryReport(file, datasetName, folderCount, sitCount, activeCount, summaryHeaders, totCounter, totPercentages, groupCount, hostCount):    
    indexHasEIF = summaryHeaders.index("hasEIF")
    indexHasCmd = summaryHeaders.index("hasCommand")
    indexActive = summaryHeaders.index("Active (to be migrated)")
    hasCmdPercent = str(round(totPercentages[indexHasCmd]*100)) + "%"
    print("FF", totCounter[indexHasEIF],  hasCmdPercent)
    
    file.write("ANALYSIS SUMMARY REPORT - " + datasetName + "\n")
    file.write("=============================================\n\n")
    file.write("Managed Systems\n")
    file.write("---------------\n\n")
    ms1 = (str(groupCount) + " Managed System Lists (MSL) found.\n")
    ms2 = (str(hostCount) + " instances of Agents found (i.e. MSL members).\n")
    file.write(ms1)
    file.write(ms2)
    file.write("\n\n")
    file.write("Situations\n")  
    file.write("----------\n\n")  
    sentence1 = (str(folderCount) + " agent folders with a total of " + str(sitCount) + " Situations of which " + str(activeCount) + " are active for migration.")
    sentence2 = ()
    file.write(sentence1)
    file.write("\n\n")
    print(sentence1)
    

def analyseMappings(folderId):
    sheet = []
    mapHeaders = ["Agent","Folder", "Supported", "Attribute Count", "Has Mapping", "No Mapping"]
    prodName = get_productCode_Name(folderId)
    totalMaps = len(metric_mappings_df)
    prodsSupported = []
    prodsNotSupported = []


    folderIds = pd.unique(metric_mappings_df["ITM_Folder"])
    folderAttributeCounts = metric_mappings_df["ITM_Folder"].value_counts()

    for f in folderIds:
        if f != "*":
            prodName = get_productCode_Name(f)
            prodSupported = is_productCode_Supported(f)    
            if prodSupported == "SUPPORTED":
                prodsSupported.append(prodName) 
            else:
                prodsNotSupported.append(prodName)
            mapRow = [prodName, f, prodSupported]
            noMappingCount = 0
            mapRow.append(folderAttributeCounts[f])
            
            folderRows = metric_mappings_df.loc[metric_mappings_df["ITM_Folder"] == f]
            for ind in folderRows.index:
                et = str(metric_mappings_df["entityType"][ind]) == "nan"
                mn = str(metric_mappings_df["metricName"][ind]) == "nan"
                q = str(metric_mappings_df["query"][ind]) == "nan"
                mappingSupported = metric_mappings_df["Supported"][ind]
                # print("MSUPPO", mappingSupported)
                if mn and q:
                    noMappingCount += 1
            mapRow.append(len(folderRows) - noMappingCount)
            mapRow.append(noMappingCount)
            sheet.append(mapRow)    
    return mapHeaders, sheet

def printSummaryList(file, summary, list):
    file.write("\n")
    file.write(summary + "\n")
    file.write("| Agent | ITM Folder ID |\n")
    file.write("| -------- | ------- |\n")
    for l in list:
        if l[1] is not None and not pd.isna(l[1]):
            file.write("| "+l[0]+" | " + str(l[1]) + "| \n")

def read_list_systems_data(filePath)-> pd.DataFrame:
    listSystems_df = pd.read_csv(filePath)
    return listSystems_df

def get_managed_systems_list_data(ls_df, group_name)-> str:
    managed_systems_list_data = None
    try:
        # FIXME remove space in generated list systems output
        managed_systems_list_csv = ls_df[ls_df['Group_Name']==group_name]['Managed_Systems_List '].item()
        managed_systems_list_data = managed_systems_list_csv.split(",")
    except:
        print("No Managed System List data found")
    return managed_systems_list_data

def get_managed_system_host_name(managed_system)-> str:
    ms = managed_system.split(":")
    print(len(ms))
    # if len(ms) > 1

def getHostNameFromManagedSystemName(distribution):
    # split based on space, then split based on : if found. 
    hostnames = []
    qManagers = []
    agent = ''
    if distribution is not None:
        dist_items = re.split('\\s|,', str(distribution))
        iterator = 0
        for iterator in range(len(dist_items)):
            dist_value = dist_items[iterator]
            hostname = ''
            # check if ":" is present in the distribution text
            if dist_value.find(':') != -1:
                managed_systems_items = dist_value.split(':')
                if len(managed_systems_items) == 2:
                    # also check for underscore - e.g. LO:UK6WPVAPSFW001_AMLBOX
                    if managed_systems_items[1].find('_') != -1:
                        managed_systems_items2 = managed_systems_items[1].split('_')
                        hostname = managed_systems_items2[0]
                        agent = managed_systems_items[0]
                    else:
                        hostname = managed_systems_items[0]
                        agent = managed_systems_items[1]
                elif len(managed_systems_items) == 3:                    
                    hostname = managed_systems_items[1]
                    agent = managed_systems_items[2]
                    if agent == 'MQ':
                        qManagers.append(managed_systems_items[1])
            else:                              
                hostname = dist_value
            if hostname !='':
                hostnames.append(hostname)
            if agent.upper().startswith('K'):
                agent = agent[1:]            
            iterator+=1
    return pd.unique(hostnames), agent, qManagers if hostnames else None

def updateManagedSystemsWithSituation(situation, groupNames, systems, listSystemsCSV):
    allGroups =  listSystemsCSV["Group_Name"]
    msl = listSystemsCSV['Managed_Systems_List ']
    for group in groupNames:
        group_index = [i for i in range(len(allGroups)) if allGroups[i] == group]
        host_names_in_group, agent, gManagers = getHostNameFromManagedSystemName(msl[group_index])     
        for host_name in host_names_in_group:
            if host_name in systems and situation not in systems[host_name].situations:
                (systems[host_name].situations).append(situation)
    return systems

# load ITM Situations root folder and analyse each folder XML files
def analyseSystemXML(inputDirectoryPath):
    logger.info("Loading ITM Managed Systems data " + inputDirectoryPath)
    listSystemsCSV = read_list_systems_data(inputDirectoryPath)
    # host_name = getHostNameFromManagedSystemName("Primary:LD1NI1900:NT,Primary:LD2NI1900:NT,Primary:SALVIN000588:NT")
    if not listSystemsCSV.empty:
        managedSystems = []
        managedSystems2 = {}
        allQqManagers = []
        # with pd.ExcelWriter(ANALYSIS_REPORT_FILENAME) as writer:
        host_names = []
        group_names = listSystemsCSV["Group_Name"]
        index = 0
        for sys in listSystemsCSV['Managed_Systems_List ']:
            host_names_in_group, agent, qManagers = getHostNameFromManagedSystemName(sys)
            if qManagers:
                allQqManagers += qManagers
            sysgroup = str(group_names[index])
            if host_names_in_group is not None:
                for host_name in host_names_in_group:
                    if host_name is not None:
                        host_names.append(host_name)                           

                        ms = ManagedSystem(
                            hostName=host_name,
                            agents=[agent],
                            situations=[],
                            groups=[sysgroup]
                        )

                        # check if the system has previously been added to the managedSystems array
                        foundSystem = next((sys for sys in managedSystems if sys.hostName == host_name), None)
                        if foundSystem:
                            # update system data, ensure no duplicate entries
                            if foundSystem.agents.count(agent) == 0:                            
                                (foundSystem.agents).append(agent)
                            if foundSystem.groups.count(sysgroup) == 0:                            
                                (foundSystem.groups).append(sysgroup)                            
                        else:
                            # create a new system entry
                            managedSystems.append(ms)  
                            managedSystems2[host_name] = ms
            index+=1
        return listSystemsCSV, group_names, host_names, managedSystems2, pd.unique(allQqManagers)
        


listSystemsCSV, group_names, host_names, systems, qManagers = analyseSystemXML(config("ITM_SYSLIST"))
analyseXML(config("ITM_FOLDER"), group_names, host_names, systems, listSystemsCSV, qManagers)
