from decouple import config
import os, time, re, json
import pandas as pd
from system import ManagedSystem, systemHeaders
from itmFolder import ITMFolder
from itmAttribute import ITMAttribute
from agent import Agent
from checklist import SituationChecklist
import numpy as np
# Import necessary spreadsheet style classes
from openpyxl.utils import get_column_letter

from preprocessor import xml_files_reader

from shared import (
    logger,
    MIGRATION_PATH,
    NAME,
    itm2instana_mappings_df,
    metric_mappings_df,
    product_codes_df,
)

from constants import (
    attribute_string_operators,
    attribute_string_keywords,
    system_event_rule,
    another_situation_reference,
    pdt_ignore,    
    INSTANA_SEV_WARNING,
    INSTANA_SEV_CRITICAL,
    INSTANA_SEV_INFO,
    SUPPORTED,
    NOT_SUPPORTED,
    CUSTOM,
    UNKNOWN,
)


# =================================
# Setup Migration Tool environment:
# =================================
# - Define Analyser constants
OUTPUT_ANALYSIS_PATH = MIGRATION_PATH + "analysis" + os.sep + NAME + os.sep
ANALYSIS_SEED_DATA_PATH = OUTPUT_ANALYSIS_PATH + "seed_data" + os.sep
ANALYSIS_REPORT_FILENAME = OUTPUT_ANALYSIS_PATH + NAME + "_report.xlsx"
ANALYSIS_SUMMARY_FILENAME = OUTPUT_ANALYSIS_PATH + NAME + "_summary.md"
ANALYSIS_SUMMARY_JSON_FILENAME = OUTPUT_ANALYSIS_PATH + NAME + "_summary.json"
ANALYSIS_SYSTEMS_JSON_FILENAME = OUTPUT_ANALYSIS_PATH + NAME + "_systems.json"
# Create folder paths if they do not already exist - makedirs succeeds even if directory exists.
os.makedirs(OUTPUT_ANALYSIS_PATH, exist_ok=True)
os.makedirs(ANALYSIS_SEED_DATA_PATH, exist_ok=True)
# Start logger
start_time = time.time()
logger.info("Analysis started...")



# =================================
# Pre-Processing
# =================================

# Create filtered list of attributes used in Situations, by folder, and check if metric mapping exists
def create_profiling_on_pdt_formula_to_attributeGroup_MetricItem(sits_list, folderId):
    msg = folderId + ": " + "Writing list of attributes found in Situations"
    logger.debug(msg)
    print(msg)
    attribute_item_list = []
    errors = []
    supported = []
    notSupported = []
    updated_sits = []
    for sit in sits_list:
        # migrate = sit.get_autostart()
        # if migrate:
            pdt_formula = sit.get_pdt()
            # split pdt formula based on spaces ' '
            list_items = pdt_formula.split()
            iterator = 0
            for iterator in range(len(list_items)):
                item = list_items[iterator]
                # find attribute_group.attribute_Item in seed data metric mappings
                if "." in item and (
                    list_items[iterator - 1] in attribute_string_keywords
                ):
                    if item not in attribute_item_list:
                        attribute_item_list.append(item)
                iterator += 1
                attribute_item_list.sort()
                # Write summary list of ITM attributes found in Situation folder data
                with open(getAnalysisSeedDataFilename(folderId), "w") as tfile:
                    tfile.write("\n".join(attribute_item_list))
    # Check if metric mapping exists
    for attrItem in attribute_item_list:
        attr = attrItem.split(".")
        metric_mapping = metric_mappings_df[
            (metric_mappings_df["ITM_Attribute_Group"] == attr[0])
            & (metric_mappings_df["ITM_Attribute_Item"] == attr[1])
        ]
        if len(metric_mapping) == 0:
            errors.append(attrItem)
        else:
            metricSupported = get_attribute_value_from_df(metric_mapping, "Supported")
            if metricSupported != "SUPPORTED":
                notSupported.append(attrItem)
            else:
                supported.append(attrItem)
    mappingMessage = str(len(attribute_item_list)) + " unique metric mappings found"
    notSupportedCount = len(np.unique(notSupported))
    supportedCount = len(np.unique(supported))
    if notSupportedCount > 0:
        isAre = "is" if notSupportedCount == 1 else "are"
        mappingMessage += (
            ", of which "
            + str(notSupportedCount)
            + " metrics "
            + isAre
            + " not supported and "
            + str(len(supported))
            + " supported."
        )
    logger.info(mappingMessage)
    for attrItem in sorted(notSupported):
        logger.warning("Metric not supported: " + folderId + " -> " + attrItem)
    for supportedMetric in sorted(supported):
        logger.info("Metric mapping found and is supported: " + supportedMetric)
    if len(errors) > 0:
        logger.info(
            str(len(errors)) + " Metric Mappings not found in folder: " + folderId
        )
        for attrItem in sorted(errors):
            logger.warning("No Metric Mapping found: " + folderId + " -> " + attrItem)
    todo = len(attribute_item_list) - notSupportedCount - supportedCount
    if todo > 0:
        logger.error("TODO!!! " + str(todo) + " metric mappings not yet mapped")


# Return seed data filename for a given folder
def getAnalysisSeedDataFilename(folderId):
    # prodSupported = is_productCode_Supported(folderId)
    return ANALYSIS_SEED_DATA_PATH + "attributeGroup_items_" + folderId + ".csv"


# Return ITM attributes list for a given folder
def loadAttributesForFolder(folderId):
    lines = None
    analysisSeedDataFilename = getAnalysisSeedDataFilename(folderId)
    logger.debug("loadAttributesForFolder:" + analysisSeedDataFilename)
    if os.path.isfile(analysisSeedDataFilename):
        try:
            with open(analysisSeedDataFilename) as file:
                lines = [line.rstrip() for line in file]
        except:
            logger.error(
                "Folder Attribute Mappings not found in seed data i.e. ",
                analysisSeedDataFilename,
            )
    return lines

def getAttributeSupportedValue(attribute, instanaSupported):
    itemgroup = attribute.split('.')
    rule_metric_mappings_df = (
                get_metricMappings_for_ITM_attributegroup_and_attributeItem(
                    attribute_group=itemgroup[0], attribute_item=itemgroup[1]
                )
            )
    metricSupported = get_attribute_value_from_df(
        rule_metric_mappings_df, "Supported"
    )
    comment = get_attribute_value_from_df(
        rule_metric_mappings_df, "comment"
    )
    if metricSupported == None and instanaSupported == "NOT_SUPPORTED":
        metricSupported = "NOT_SUPPORTED"
    elif metricSupported == "NOT_SUPPORTED":
        roadmap = get_attribute_value_from_df(
            rule_metric_mappings_df, "Roadmap"
        )
        if roadmap != None:
            metricSupported = "ROADMAP"
            comment = roadmap
    
    return metricSupported, comment,
    

def getAttributes(attributeList, supported):
    attributes = []
    supportedCounts = []
    comments = []
    for attr in attributeList:
        metricSupported, comment = getAttributeSupportedValue(attr, supported)
        itemgroup = attr.split('.')            
        attribute =  ITMAttribute (
            attrItem=itemgroup[0],
            attrValue=itemgroup[1],
            supported=metricSupported,
            comment=comment
        )
        attributes.append(attribute)
        supportedCounts.append(metricSupported)
        if comment != None:
            comments.append(comment)
    return attributes, supportedCounts, pd.unique(comments)

# Return name associated with Product code using ITMProductCodes.csv
def is_productCode_Supported(productCode):
    supported = None
    # special case for Local Time attributes which may be associated with multiple products
    if productCode == "*": supported = "SUPPORTED"  
    try:
        supported = product_codes_df[(product_codes_df["ProductCode"] == productCode)][
            "Supported"
        ].item()
    except:
        if not isNaN(productCode):
            logger.warning("Agent Folder not recognised: " + productCode)
    return supported

# Return name associated with Product code using ITMProductCodes.csv
def get_productCode_Name(productCode):
    productCodeName = None
    try:
        productCodeName = product_codes_df[
            (product_codes_df["ProductCode"] == productCode)
        ]["ProductName"].item()
    except:
        if not isNaN(productCode):
            logger.warning("Agent Folder not recognised: " + productCode)
    if isNaN(productCodeName):
        productCodeName = ''
    return productCodeName

def preprocessData(inputDirectoryPath):
    logger.info("Loading ITM Situations found in " + inputDirectoryPath)
    itmFolders = {}    
    allSituationReferences = []
    allAgents={}
    if os.path.isdir(inputDirectoryPath) and os.path.exists(inputDirectoryPath):
        for folder, subs, files in os.walk(inputDirectoryPath):                
            subFolders = os.path.split(folder)
            folderId = subFolders[len(subFolders) - 1]
            if folderId not in itmFolders:
                if len(files) > 0 and len(subs) == 0:
                    # Load ITM Situations by folder
                    sits_list = xml_files_reader(folder)
                    create_profiling_on_pdt_formula_to_attributeGroup_MetricItem(
                        sits_list, folderId
                    )
                    prodName = get_productCode_Name(folderId)
                    supported = is_productCode_Supported(folderId)
                    attributeList = loadAttributesForFolder(folderId)    
                    attributes, supportedCounts, comments = getAttributes(attributeList, supported)
                    folderSituationReferences = getAllSituationReferences(folderId, sits_list)
                    if len(folderSituationReferences) > 0:
                        allSituationReferences.extend(folderSituationReferences)
                    
                    updatedSituations = checkSituations(sits_list, folderId, attributes, allSituationReferences)
                    # for sit in updatedSituations:
                    #     print("updatedSituations::" + updatedSituations[sit].sitName + " migrate=" + str(updatedSituations[sit].migrate) + " until=" + str(updatedSituations[sit].untilRef))
                    supportedPct = supportedCounts.count("SUPPORTED")/len(attributeList)
                        
                    agent = Agent(
                        productCode=folderId,
                        name=prodName,
                        supported=supported,
                        attributeSupportedCount=supportedCounts.count("SUPPORTED"),
                        attributeSupportedCountPercentage=toPercentage(supportedPct),
                        attributeNotSupportedCount=supportedCounts.count("NOT_SUPPORTED"),
                        attributeRoadmapCount=supportedCounts.count("ROADMAP"),
                        attributeOtherCount=supportedCounts.count("TODO") + supportedCounts.count(None),
                        attributeTotalCount=len(attributeList),
                        comments=comments
                    )

                    allAgents[folderId] = agent
                    itmFolder = ITMFolder(
                        folderId=folderId,
                        supported=supported,
                        situations=updatedSituations,                         
                        sitRefs=folderSituationReferences,
                        description=prodName,
                        attributes=attributes
                    ) 
                    itmFolders[folderId] = itmFolder
            else:
                print("Folder not in itmFolders", folderId)
    return itmFolders, updatedSituations, allAgents

def getAllSituationReferences(folderId, situations):
    sitRefs = []
    for sit in situations:
        pdt_formula = sit.get_pdt()
        sitRefName = searchForUntilSitRefInPDT(pdt_formula)
        if sitRefName != None and sitRefName not in sitRefs:
            sitRefs.append(sitRefName)
    return sitRefs

def searchForUntilSitRefInPDT(pdt_formula):
    # split pdt formula based on spaces ' '
    pdt_array = pdt_formula.split()
    if another_situation_reference in pdt_array:            
        sitRefIndex = pdt_array.index(another_situation_reference)
        return pdt_array[sitRefIndex+1]

def checkSituations(itmSituations, folderId, attributes, allSituationReferences):
    folderSituations = {}
    for situation in itmSituations:
        sitRefName = searchForUntilSitRefInPDT(situation.pdt)
        pdtSupported = analyseSituationPDT(situation.pdt, folderId, attributes)        
        attributesSupportedList = pdtSupported['MappingConfidence']
        pdtSupported = True
        for entry in attributesSupportedList:
            if entry != "SUPPORTED":
                pdtSupported = False
        # if pdtSupported:
            # print("WOWOWWOWOWOWO")
        sit = SituationChecklist(
            sitName=situation.sit_name,
            folderId=folderId,
            autoStart=situation.get_autostart(),
            pdt=situation.pdt,
            pdtMappings=attributesSupportedList,
            severity=situation.get_severity(),
            isRef=situation.sit_name in allSituationReferences,
            untilRef=sitRefName,
            itmSituation=situation,
            migrate=pdtSupported
        )
        folderSituations[situation.sit_name] = sit
        sev = situation.get_severity() 
    # if folderId == "MQ":
    #     print("!!!MQ!!!")
    #     for sit in folderSituations:
    #         print(folderSituations[sit].sitName + " migrate=" + str(folderSituations[sit].migrate) + " until=" + str(folderSituations[sit].untilRef))
    return folderSituations

def getSeverity(situations):
    sevInfos = []
    sevCrits = []
    sevWarns = []
    sevOthers = []
    for situation in situations:
        sev = situation.get_severity()            
        sevMap = get_keyword_mapping(sev)
        sevCritical=sevMap == INSTANA_SEV_CRITICAL
        sevWarning=sevMap == INSTANA_SEV_WARNING
        sevInfo=sevMap == INSTANA_SEV_INFO
        sevOther=sevMap != INSTANA_SEV_CRITICAL and sevMap != INSTANA_SEV_WARNING and sevMap != INSTANA_SEV_INFO

# =================================
# Spreadsheet report utilities
# =================================
def isNaN(num):
    return num != num

def read_list_systems_data(filePath)-> pd.DataFrame:
    listSystems_df = pd.read_csv(filePath)
    return listSystems_df


def get_keyword_mapping(itm_keyword):
    keyword_mapping = None
    if itm_keyword != None:
        try:
            # TODO: pass dynamic columns names from dataframe rather than hard coded values
            keyword_mapping = itm2instana_mappings_df[
                itm2instana_mappings_df["ITM_XML"] == itm_keyword
            ]["Instana_JSON_value"].item()
        except:
            logger.warning(
                "Keyword Mappings not found in seed data i.e. itm2instanaMapping.csv : "
                + str(itm_keyword)
            )
    return keyword_mapping

def get_attribute_value_from_df(attribute_df, attribute_name):    
    attribute_value = None
    try:
        # set to None, if record is present but value is empty in dataframe
        if attribute_df[attribute_name].item() and not isNaN(
            attribute_df[attribute_name].item()
        ):
            attribute_value = attribute_df[attribute_name].item()
    except:
        attribute_value = None    
    return attribute_value


def get_metricMappings_for_ITM_attributegroup_and_attributeItem(
    attribute_group, attribute_item
):
    metric_mapping = None
    try:
        metric_mapping = metric_mappings_df[
            (metric_mappings_df["ITM_Attribute_Group"] == attribute_group)
            & (metric_mappings_df["ITM_Attribute_Item"] == attribute_item)
        ]
    except:
        logger.warning("Metric Mappings not found in seed data i.e. metricMapping.csv")
    return metric_mapping

      

def analyseSituationPDT(pdt_formula, itmFolder, attributes):
    rules = []  # rules object to be added to the current event
    rangeMetrics = []
    supported = []
    missingMetric = []
    query = None
    entity_type = None
    mappingConfidence = []
    sitReport = {}    

    pdt_array = pdt_formula.split()
    # if another_situation_reference in pdt_array:
    #     logger.debug(
    #         "The situation referencing another situation not supported. " + pdt_formula
    #     )
    #     sitRefIndex = pdt_array.index(another_situation_reference)
    #     if pdt_array[sitRefIndex+1] in itmFolder.refs:
    #         mappingConfidence.append('SUPPORTED')
    #     print(pdt_array[sitRefIndex+1])
        

    # check if pdt formula items are found in system event rule array
    is_system_rule = [True for item in pdt_array if item in system_event_rule]
    # sitReport.isSystem = len(is_system_rule) > 0
    # print("LOUISE attrinbute", attributes)
    def get_values_for_single_condition(list_items_in_single_pdt):
        iterator = 0
        count = 0
        hasSitRef = False
        rule_metric_mappings_df = None
        condition_operator = None
        operand = None
        query_value = None
        attr = ''
        for iterator in range(len(list_items_in_single_pdt)):
            item = list_items_in_single_pdt[iterator]
            # *IF *VALUE Queue_Statistics.Queue_Name *EQ Q289P.PROD.SA.SAST.REQUEST *AND *VALUE Queue_Statistics.Current_Depth *GE 1000
            #
            if item == "*IF" or item in attribute_string_keywords:
                keyword_string = item
                continue  # skipping

            # next item should be the situation name after hasSitRef is set
            if hasSitRef == True:
                # check if item has full mapping FIXME LOUISE
                mappingConfidence.append("SUPPORTED")
                hasSitRef = False
                continue
            
            if item == "*UNTIL" or item == "*SIT":
                hasSitRef = True
                continue
            
            # find attribute_group.attribute_Item, then split and retrieve metric mappings from seed data
            if "." in item and (                
                list_items_in_single_pdt[iterator - 1] in attribute_string_keywords
            ):
                attr = item
                items = item.split(".")
                metricSupported = None
                for attribute in attributes:
                    if attribute.attrItem == items[0] and attribute.attrValue == items[1]:
                        metricSupported = attribute.supported
                        if metricSupported is not None:            
                            mappingConfidence.append(metricSupported)

                rule_metric_mappings_df = (
                    get_metricMappings_for_ITM_attributegroup_and_attributeItem(
                        attribute_group=items[0], attribute_item=items[1]
                    )
                )

            # check if there is an *EQ keyword, fetch its mapping value from seed data
            elif item in attribute_string_operators:
                condition_operator = get_keyword_mapping(item)
            elif (                                
                item 
                and list_items_in_single_pdt[iterator - 1] in attribute_string_operators
            ):
                operand = item.strip("'\"")

            iterator += 1        
        
        entity_type = get_attribute_value_from_df(rule_metric_mappings_df, "entityType")
        metric_name = get_attribute_value_from_df(rule_metric_mappings_df, "metricName")
        aggregation = get_attribute_value_from_df(
            rule_metric_mappings_df, "aggregation"
        )
        query_value = get_attribute_value_from_df(rule_metric_mappings_df, "query")
        metric_pattern_prefix = get_attribute_value_from_df(
            rule_metric_mappings_df, "metricPattern.prefix"
        )
        metric_pattern_postfix = get_attribute_value_from_df(
            rule_metric_mappings_df, "metricPattern.postfix"
        )

        # if metric_pattern_prefix or metric_pattern_postfix and sitReport.hasMetricPattern != True:
        # sitReport.hasMetricPattern = (
        #     metric_pattern_prefix != None or metric_pattern_postfix != None
        # )
        
        conversion_rule = get_attribute_value_from_df(
            rule_metric_mappings_df, "conversionRule"
        )
        
        # metricSupported = get_attribute_value_from_df(
        #     rule_metric_mappings_df, "Supported"
        # )
        
        # if metricSupported is not None:            
        #     mappingConfidence.append(metricSupported)
        
        # sitReport.hasMaintWindow = conversion_rule == maint_window
        
        return (
            entity_type,
            metric_name,
            metric_pattern_prefix,
            metric_pattern_postfix,
            aggregation,
            query_value,
            condition_operator,
            operand,
            mappingConfidence,
            conversion_rule
        )

    def handle_single_condition(condition, query_operator, query, rules, entity_type):
        logger.debug(condition)
        # split pdt formula based on spaces ' '
        list_items_in_single_pdt = condition.split()
        (
            rule_entity_type,
            metric_name,
            metric_pattern_prefix,
            metric_pattern_postfix,
            aggregation,
            query_value,
            condition_operator,
            operand,
            metricSupported,
            conversion_rule
        ) = get_values_for_single_condition(list_items_in_single_pdt)

        supported.append(metricSupported)

        if entity_type is None:
            entity_type = rule_entity_type
        if (
            (
                metric_name is None
                and (metric_pattern_prefix is not None or query_value is not None)
            )
            or (
                metric_pattern_prefix is None
                and (metric_name is not None or query_value is not None)
            )
            or (
                query_value is None
                and (metric_name is not None or metric_pattern_prefix is not None)
            )
        ):
            sitReport["hasMapping"] = True
        else:
            sitReport["hasMapping"] = False

        if (
            rule_entity_type is not None
            and metric_name is None
            and metric_pattern_prefix is None
            and metric_pattern_postfix is None
            and query_value is not None
        ):
            if query is None:
                if operand is not None:
                    query = query_value + ":'" + operand + "'"
                # else:
                    # print("LOUISE QUERRRY", pdt_formula)
            elif query is not None:
                if operand is not None and query_operator is not None:
                    query = query + query_operator + query_value + ":'" + operand + "'"
        elif rule_entity_type is not None and metric_name is not None:
            duplicate_metric_rule = [r for r in rules if r["metricName"] == metric_name]
            if (
                duplicate_metric_rule
                and duplicate_metric_rule[0].get("conditionOperator").startswith(">")
                and condition_operator.startswith("<")
            ) or (
                duplicate_metric_rule
                and duplicate_metric_rule[0].get("conditionOperator").startswith("<")
                and condition_operator.startswith(">")
            ):
                rangeMetrics.append(metric_name)
                logger.debug(
                    "A RANGE with a duplicate metricName was found in PDT formula. Skipped adding a new rule into rules array."
                )
            else:
                rule = {
                    "ruleType": "threshold",
                    "metricName": metric_name,
                    "metricPattern": None,
                    "rollup": 0,
                    "window": 1000,
                    "aggregation": aggregation,
                    "conditionOperator": condition_operator,
                    "conditionValue": operand,
                }
                rules.append(rule)
        return query, rules, entity_type

    def handle_conditions(
        pdt_formula,
        query,
        rules,
        entity_type,
        query_operator=" OR ",
    ):
        if pdt_formula not in pdt_ignore and pdt_formula:
            pdt_sub_array = pdt_formula.split()
            if "(" in pdt_sub_array or ")" in pdt_sub_array:
                logger.debug("handling brackets, splitting ")
                pdt_regex_array = re.split(r"\(|\)", pdt_formula)
                logger.debug(pdt_regex_array)
                for item in pdt_regex_array:
                    query, rules, entity_type = handle_conditions(
                        item.strip(), query, rules, entity_type
                    )
            elif "*OR" in pdt_sub_array:
                query_operator = " OR "
                logger.debug("handling *ORs")
                pdt_ors = pdt_formula.split("*OR")
                logger.debug(pdt_ors)
                for item in pdt_ors:
                    query, rules, entity_type = handle_conditions(
                        item.strip(), query, rules, entity_type
                    )
            elif "*AND" in pdt_sub_array:
                query_operator = " AND "
                logger.debug("Handling *ANDs")
                pdt_ands = pdt_formula.split("*AND")
                logger.debug(pdt_ands)
                for item in pdt_ands:
                    query, rules, entity_type = handle_conditions(
                        item.strip(), query, rules, entity_type
                    )
            else:
                logger.debug("handling single condition() " + pdt_formula)
                query, rules, entity_type = handle_single_condition(
                    pdt_formula, query_operator, query, rules, entity_type
                )

        return query, rules, entity_type

    query, rules, entity_type = handle_conditions(
        pdt_formula,
        query,
        rules,
        entity_type,
    )

    sitReport["MappingConfidence"] = mappingConfidence
    # sitReport.hasRange = len(rangeMetrics) > 0
    # sitReport.supported = supported.count("SUPPORTED")
    # sitReport.notSupported = supported.count("NOT_SUPPORTED")
    # sitReport.todo = supported.count("TODO")
    # sitReport.missing = len(missingMetric)
    # sitReport.mappingScore = len(mappingConfidence) - mappingConfidence.count("SUPPORTED")
    return sitReport




# load ITM Situations root folder and analyse each folder XML files
def analyseSystemXML(inputDirectoryPath, itmFolders, updatedSituations):
    logger.info("Loading MUFG ITM Managed Systems data " + inputDirectoryPath)
    listSystemsCSV = read_list_systems_data(inputDirectoryPath)
    
    if not listSystemsCSV.empty:
        managedSystems = {}
        host_names = listSystemsCSV["Host Name"]
        group_names = listSystemsCSV["Managed System List"]
        agents = listSystemsCSV["Product Code"]
        system_situations = listSystemsCSV["Situation Name"]
        formulas = listSystemsCSV["Formula"]
        index = 0
        
        # reportSituations = {}
        for sys in host_names:  
            productCode = agents[index]    
            prodSupported = is_productCode_Supported(productCode)
            sitName = system_situations[index]   

            # migrateYes = 0
            # migrateNo = 0
            
            
            # if prodSupported and productCode in itmFolders:           
            # # if productCode in itmFolders:                                    
            #     folderSituations = itmFolders[productCode].situations
            #     if sitName in folderSituations:
            #         situation = folderSituations[sitName]
            #         if situation.migrate:
            #             migrateYes += 1
            #         else:
            #             migrateNo += 1


                    # print(situation.migrate)
                    # if sitName not in migrate:
                    #     migrate.append(sitName)
                    # systemSituation = SituationChecklist(
                    #     sitName=situation.sitName,
                    #     pdt=situation.pdt                  
                    # )
                    # reportSituations[sitName] = systemSituation
                    # reportSituations[sitName] = folderSituations[sitName] 

                            
                # f = analyseSituationPDT(formulas[index], itmFolders[productCode])
                # conf = f["MappingConfidence"]
                # if len(conf) > 0 and conf.count("SUPPORTED") == len(conf): 
                #     confidence = 1   
            fullSitName = productCode + ":" + sitName
            # check if the system has previously been added to the managedSystems array
            if sys not in managedSystems:
                ms = ManagedSystem(
                    hostName=sys,
                    agents=[productCode],
                    situations=[fullSitName],
                    groups=[group_names[index]],
                    # migrateYes=migrateYes,
                    # migrateNo=migrateNo
                )
                managedSystems[sys] = ms
            else:
                foundSystem = managedSystems[sys]
                # update system data, ensure no duplicate entries
                if (foundSystem.agents.count(productCode)) == 0:                            
                    (foundSystem.agents).append(str(productCode))
                if foundSystem.groups.count(group_names[index]) == 0:                            
                    (foundSystem.groups).append(group_names[index])  
                if foundSystem.situations.count(fullSitName) == 0:           
                    (foundSystem.situations).append(fullSitName)  
                foundSystem.situationCount = len(foundSystem.situations)
                # foundSystem.migrateYes += migrateYes
                # foundSystem.migrateNo += migrateNo
             
            index+=1
        for sys in managedSystems:
            systemSituations = managedSystems[sys].situations
            migrateYes = 0
            migrateNo = 0
            roadmap = 0
            for syssit in systemSituations:
                ss = syssit.split(":")
                prodCode = str(ss[0]) if ss[0].isdigit() else ss[0]
                if prodCode.isdigit() and int(prodCode) < 10:
                    prodCode = "0" + prodCode
                
                folderSituations = itmFolders[prodCode].situations if prodCode in itmFolders else None           
                ssitName = ss[1]
                # LOUISE FIXME
                if folderSituations != None and ssitName in folderSituations:
                    situation = folderSituations[ssitName]
                    if situation.migrate:
                        migrateYes += 1
                    else:
                        migrateNo += 1
                        if not situation.pdtMappings.count("NOT_SUPPORTED") > 0:
                            roadmap += 1
            managedSystems[sys].migrateYes = migrateYes
            managedSystems[sys].migrateNo = migrateNo
            managedSystems[sys].roadmap = roadmap
        updatedSystems = managedSystems
        # updateManagedSystems(managedSystems, itmFolders)
        return listSystemsCSV, pd.unique(group_names), pd.unique(host_names), updatedSystems

def canMigrate(situation, itmFolders):
    sit = situation.split(":")
    productCode = sit[0]
    sitName = sit[1]
    if productCode == "MQ":
        folderSituations = itmFolders[productCode].situations
        if sitName in folderSituations:
            situation = folderSituations[sitName]
            print(situation.pdt)
    return False

def updateManagedSystems(managedSystems, itmFolders):
    updatedSystems = {}
    
    for sys in managedSystems:
        updatedSystem = managedSystems[sys]
        migrateYes = []
        migrateNo = []
        for sit in updatedSystem.situations:
            if canMigrate(sit, itmFolders):
                migrateYes.append(sit)
            else: 
                migrateNo.append(sit)
        updatedSystem.migrateYes = len(migrateYes)
        updatedSystem.migrateNo = len(migrateNo)
        updatedSystems[sys] = updatedSystem

    return updatedSystems

# Create short summary paragraph of analysis results, e.g.:
# Client data has 22 folders with a total of 858 Situations of which 503 are active for migration. 
# 
def writeSummaryReport(file, datasetName, folderCount, sitCount, activeCount, summaryHeaders, totCounter, totPercentages, groupCount, hostCount):    
    file.write("ANALYSIS SUMMARY REPORT - " + datasetName + "\n")
    file.write("=============================================\n\n")
    file.write("Filename: " + datasetName)
    file.write("\n\n")
    file.write("Managed Systems\n")
    file.write("---------------\n\n")
    ms1 = (str(groupCount) + " Managed System Lists (MSL) found.\n")
    ms2 = (str(hostCount) + " instances of Agents found (i.e. MSL members).\n")
    file.write(ms1)
    file.write(ms2)
    file.write("\n\n")
    file.write("Situations\n")  
    file.write("----------\n\n")  
    sentence1 = (str(folderCount) + " agent folders with a total of " + str(sitCount) + " Situations of which " + str(activeCount) + " are active for migration.")
    sentence2 = ()
    file.write(sentence1)
    file.write("\n\n")
    print(sentence1)
   
def printSummaryList(file, summary, list):
    file.write("\n")
    file.write(summary + "\n")
    file.write("| Agent | ITM Folder ID |\n")
    file.write("| -------- | ------- |\n")
    for l in list:
        if l[1] is not None and not pd.isna(l[1]):
            file.write("| "+l[0]+" | " + str(l[1]) + "| \n")


def concatArrayAsString(arr):
    return ','.join(map(str, arr)) if len(arr) > 0 else ''             
  
def analyseXML(inputDirectoryPath, group_names, host_names, systems, listSystemsCSV):
    logger.info("Loading ITM Situations found in " + inputDirectoryPath)
    iterator = 0
    if os.path.isdir(inputDirectoryPath) and os.path.exists(inputDirectoryPath):
        with pd.ExcelWriter(ANALYSIS_REPORT_FILENAME) as writer:            
            summaryDFs = []
            dataDFs = {}
            prodsSupported = []
            prodsNotSupported = []
            systemSituations = systems
            folderAttributes = {}

            # Managed Systems data tabs
            if len(group_names) > 0: 
                pd.DataFrame(group_names).to_excel(
                        writer, sheet_name="Managed System List", header=["MSL GroupName"], index=None,
                    ) 
            hostsystems = []
            for host in host_names:
                sys = systemSituations[host]
                typeList = concatArrayAsString(sys.types)
                agentList = concatArrayAsString(sys.agents)
                groupList = concatArrayAsString(sys.groups)
                situationList = concatArrayAsString(sys.situations)
                hostsystem = [host, len(sys.types), typeList, len(sys.agents), agentList, len(sys.groups), groupList, len(sys.situations), situationList]
                hostsystems.append(hostsystem)

            if len(hostsystems) > 0: 
                pd.DataFrame(hostsystems).to_excel(
                        writer, sheet_name="Systems", header=systemHeaders, index=None, freeze_panes=(1, 1),
                    )

            logger.info("Analysis report saved as " + ANALYSIS_REPORT_FILENAME)
            
            with open(ANALYSIS_SUMMARY_FILENAME, "w") as summaryFile:
                writeSummaryReport(summaryFile, inputDirectoryPath, listSystemsCSV.size, 0, 0, 0, 0, 0, len(group_names), len(host_names))
                sentence1 = "The Migration Tool may be able to migrate attributes from the following " + str(len(prodsSupported)) + " ITM agent types, "
                sentence2 = "(i.e. some but not all ITM attributes have mappings):\n"
                printSummaryList(summaryFile, sentence1 + sentence2, prodsSupported)
                
                printSummaryList(summaryFile, "The following " + str(len(prodsNotSupported)) + " ITM agent types are not supported by the Migration Tool:", prodsNotSupported)
                summaryFile.close()           

    else:
        logger.error(
            "Please provide a valid directory path in 'ITM_FOLDER' environment variable."
        )
        return None
    logger.info(
        "Total execution time: " + str(round(time.time() - start_time, 3)) + " seconds"
    )
    return iterator

def toPercentage(amount):
    return str(round(amount*100)) + "%"

def writeToJSON(data):    
    out = json.dumps(data, default=vars, sort_keys=True, indent=4, ensure_ascii=False)
    return out


def getSystemSummary(systems):       
    allAgents = []
    for sys in systems:
        if systems[sys].agents != None:
            for agent in systems[sys].agents:
                if agent not in allAgents:
                    allAgents.append(agent)
    totalAgentsCount = len(allAgents)    
    print(str(len(systems)) + " Resources in the estate are managed by IBM Tivoli Monitoring across " + str(totalAgentsCount) + " agent types")

def writeSheet(writer, sheetName, headers, data, freezeCol, firstColumnWidth, isSummarySheet):
    if freezeCol is None:
        freezeCol = 1
    if len(data) > 0:         
        pd.DataFrame(data)  .to_excel(
            writer, sheet_name=sheetName, header=headers, index=None, freeze_panes=(1, freezeCol), 
        )
    # Format Summary sheet cell values
    summarySheet = writer.sheets[sheetName]
    formatSummarySheet(summarySheet, firstColumnWidth, isSummarySheet)


def formatSummarySheet(summarySheet, firstColumnWidth, isSummarySheet):
    # cell ranges
    title_row = "1"
    total_row = summarySheet.max_row - 1
    percent_row = summarySheet.max_row
    value_cells = "B2:{col}{row}".format(
        col=get_column_letter(summarySheet.max_column), row=summarySheet.max_row
    )
    index_column = "A"
    active_rows = "E2:E{row}".format(
        # col=get_column_letter(summarySheet.max_column),
        row=summarySheet.max_row
        - 2
    )

    # set first column width (Agent name)
    if firstColumnWidth != None:
        summarySheet.column_dimensions[index_column].width = firstColumnWidth

    for row in summarySheet[value_cells]:
        for cell in row:
            if cell.value == SUPPORTED:
                cell.style = "Good"
            elif cell.value == NOT_SUPPORTED or cell.value == UNKNOWN:
                cell.style = "Bad"
            elif cell.value == CUSTOM:
                cell.style = "Neutral"
            else:
                cell.style = "Normal"
    # for row in summarySheet[active_rows]:
    #     for cell in row:
    #         if cell.value == 0:  
    #             cell.style = "Check Cell"
    # style header line last, so that headline style wins in cell A1
    if isSummarySheet:
        for cell in summarySheet[title_row]:
            cell.style = "Headline 2"
        for cell in summarySheet[total_row]:
            cell.style = "Total"
        for cell in summarySheet[percent_row]:
            cell.style = "Percent"

def getAttributesFromITMFolders():
    attributes = []
    for itmFolder in itmFolders:     
        folder = itmFolders[itmFolder]   
        itmAttributes = folder.attributes
        for attribute in itmAttributes:
            itmAttribute = [folder.folderId, attribute.attrItem, attribute.attrValue, attribute.supported, attribute.comment]
            attributes.append(itmAttribute)
    sheetName = "Attributes"
    headers = ["Folder","Attribute Group", "Attribute Item", "Supported", "Comment"]
    freezeCol = 1
    firstColumnWidth = 10
    return (sheetName, headers, attributes, freezeCol, firstColumnWidth)

def getSituationsFromITMFolders():
    data = []
    for itmFolder in itmFolders:     
        folder = itmFolders[itmFolder]   
        itmSituations = folder.situations
        for situation in itmSituations:
            sit = itmSituations[situation]
            itmSituation = [sit.sitName, folder.folderId, sit.migrate, sit.autoStart, sit.severity, sit.isRef, sit.untilRef, sit.pdt, sit.pdtMappings, sit.distribution]
            data.append(itmSituation)
    sheetName = "Situations"
    headers = ["Situation", "Folder", "Migrate?", "autoStart", "Severity", "IsRef", "UntilRef", "Formula", "Mappings", "Distribution"]
    freezeCol = 1
    firstColumnWidth = 30
    return (sheetName, headers, data, freezeCol, firstColumnWidth)

def getAgents(allAgents):
    data = []
    for folderId in allAgents:     
        agent = allAgents[folderId]   
        agentSummary = [agent.name, agent.productCode, agent.supported, agent.attributeSupportedCount, agent.attributeSupportedCountPercentage, agent.attributeNotSupportedCount, agent.attributeRoadmapCount, agent.attributeOtherCount, agent.attributeTotalCount, concatArrayAsString(agent.comments)]
        data.append(agentSummary)
    sheetName = "Agents"
    headers = ["Agent", "Id", "Instana Supported", "Supported Attributes", "%", "Not supported", "Roadmap", "Other", "Total", "Comments"]
    freezeCol = 1
    firstColumnWidth = 30
    return (sheetName, headers, data, freezeCol, firstColumnWidth)

# first pre-process to get situation references
print("Preprocessing...")
itmFolders, updatedSituations, allAgents = preprocessData(config("ITM_FOLDER"))
mqRefs = itmFolders["MQ"]

with open(ANALYSIS_SUMMARY_JSON_FILENAME, "w") as outfile: 
    # for productCode in itmFolders:
    #     outfile.write(writeToJSON(itmFolders[productCode]))
    outfile.write(writeToJSON(itmFolders))
    outfile.close()

listSystemsCSV, group_names, host_names, systems = analyseSystemXML(config("MUFG_SYSLIST"), itmFolders, updatedSituations)

# print(systems["GBLVAP001569"].agents)
# print(systems["GBLVAP001569"].situations)
# print(systems["GBLVAP001569"].mappings)
# print(systems["GBLVAP001569"].groups)

getSystemSummary(systems)
systems["GBLVAP001569"].getManagedSystemReportArray()

with open(ANALYSIS_SYSTEMS_JSON_FILENAME, "w") as outfile: 
    outfile.write("{\n")
    count = 0
    for sys in systems:
        if not isNaN(sys):
            outfile.write('\"' + sys + '\": ')
            outfile.write(writeToJSON(systems[sys]))
            if count < len(systems) - 2:
                outfile.write(",")
        count+=1
    outfile.write("\n}")
    outfile.close()


hostsystems = []
for host in systems:
    sys = systems[host]
    agentList = ','.join(map(str, sys.agents)) if len(sys.agents) > 0 else ''
    groupList = ','.join(map(str, sys.groups)) if len(sys.groups) > 0 else ''
    situationCount = sys.situationCount
    migrateYes = sys.migrateYes
    migrateNo = sys.migrateNo
    roadmap = sys.roadmap
    hostsystem = [host, len(sys.agents), agentList, len(sys.groups), groupList, situationCount, migrateYes, migrateNo, roadmap]                
    hostsystems.append(hostsystem)

with pd.ExcelWriter(ANALYSIS_REPORT_FILENAME) as writer: 
    if len(hostsystems) > 0: 
        pd.DataFrame(hostsystems).to_excel(
                        writer, sheet_name="Systems", header=["HostName","Agent count", "Agents", "Group count", "Groups",  "SitCount","MigrateYes" , "MigrateNo", "Roadmap"], index=None, freeze_panes=(1, 1),
                    )
    sheetName, headers, attributes, freezeCol, firstColumnWidth = getSituationsFromITMFolders()
    writeSheet(writer,sheetName, headers, attributes, freezeCol, firstColumnWidth, False )
    sheetName, headers, attributes, freezeCol, firstColumnWidth = getAttributesFromITMFolders()
    writeSheet(writer,sheetName, headers, attributes, freezeCol, firstColumnWidth, False )
    sheetName, headers, attributes, freezeCol, firstColumnWidth = getAgents(allAgents)
    writeSheet(writer,sheetName, headers, attributes, freezeCol, firstColumnWidth, False )
    
    



# listSystemsCSV, group_names, host_names, systems = analyseSystemXML(config("MUFG_SYSLIST"), itmFolders)
# with open(ANALYSIS_SYSTEMS_JSON_FILENAME, "w") as outfile: 
#     outfile.write(writeToJSON(systems.getManagedSystemReportArray))
#     outfile.close()

# analyseXML(config("ITM_FOLDER"), group_names, host_names, systems, listSystemsCSV)
